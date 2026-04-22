"""
خدمة تصدير التوزيعات إلى Excel.
"""

from __future__ import annotations

import io
from itertools import groupby
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from distributions.selectors import get_assignments_for_export


class ExportService:
    """يُصدّر بيانات التوزيع إلى ملف Excel."""

    HEADER_FILL = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    CELL_FONT   = Font(name="Arial", size=11)
    CENTER      = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ألوان متناوبة بين القاعات
    ROOM_FILLS = [
        PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),  # أزرق فاتح
        PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),  # أخضر فاتح
    ]

    HEADERS = ["اسم المراقب", "القاعة", "مكان القاعة", "التاريخ", "الوقت"]

    # ── حدود خارجية للخلايا المدمجة ─────────────────────────────
    @staticmethod
    def _outer_border(top: bool = True, bottom: bool = True,
                      left: bool = True, right: bool = True) -> Border:
        thin = Side(style="thin")
        none = Side(style=None)
        return Border(
            top=thin if top else none,
            bottom=thin if bottom else none,
            left=thin if left else none,
            right=thin if right else none,
        )

    @classmethod
    def export(cls, batch_id: int | None = None) -> io.BytesIO:
        """
        يُنشئ ملف Excel بـ:
          - خلايا مدمجة للقاعة / التاريخ / الوقت لكل مجموعة مراقبين
          - تلوين متناوب بين كل قاعة والأخرى
        """
        assignments = list(get_assignments_for_export(batch_id))

        wb = Workbook()
        ws = wb.active
        ws.title = "التوزيعات"
        ws.sheet_view.rightToLeft = True

        # ── صف العناوين ─────────────────────────────────────────
        for col_idx, header in enumerate(cls.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = cls.HEADER_FONT
            cell.fill      = cls.HEADER_FILL
            cell.alignment = cls.CENTER
            cell.border    = cls.THIN_BORDER
        ws.row_dimensions[1].height = 26

        if not assignments:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        # ── التجميع حسب (التاريخ، رقم القاعة) ─────────────────
        current_row = 2
        color_idx   = 0

        for (date, room_label), group_iter in groupby(
            assignments, key=lambda a: (a.date, a.room_label)
        ):
            group_items = list(group_iter)
            group_start = current_row
            fill        = cls.ROOM_FILLS[color_idx % len(cls.ROOM_FILLS)]
            time_val    = group_items[0].time if group_items else ""

            # ── كتابة صفوف المراقبين ────────────────────────────
            # استخراج مكان القاعة
            location_val = ""
            if group_items and group_items[0].classroom:
                location_val = getattr(group_items[0].classroom, "location", "") or ""

            for slot_idx, item in enumerate(group_items):
                display_name = (
                    item.teacher.formatted_name if item.teacher
                    else item.teacher_name
                )

                # عمود A: اسم المراقب (كل صف)
                ca = ws.cell(row=current_row, column=1, value=display_name)
                ca.font      = cls.CELL_FONT
                ca.fill      = fill
                ca.alignment = cls.CENTER
                ca.border    = cls.THIN_BORDER

                # أعمدة B, C, D, E: نكتب القيمة في الصف الأول فقط
                if slot_idx == 0:
                    for col_idx, val in enumerate(
                        [room_label, location_val, date, time_val], start=2
                    ):
                        c = ws.cell(row=current_row, column=col_idx, value=val)
                        c.font      = cls.CELL_FONT
                        c.fill      = fill
                        c.alignment = cls.CENTER
                        c.border    = cls.THIN_BORDER

                ws.row_dimensions[current_row].height = 22
                current_row += 1

            group_end = current_row - 1

            # ── دمج خلايا القاعة / المكان / التاريخ / الوقت ────
            if group_end > group_start:
                for col_letter in ("B", "C", "D", "E"):
                    ws.merge_cells(
                        f"{col_letter}{group_start}:{col_letter}{group_end}"
                    )
                    # إعادة تطبيق التنسيق على الخلية المدمجة
                    mc = ws[f"{col_letter}{group_start}"]
                    mc.fill      = fill
                    mc.font      = cls.CELL_FONT
                    mc.alignment = cls.CENTER
                    mc.border    = cls.THIN_BORDER

            color_idx += 1

        # ── عرض الأعمدة ─────────────────────────────────────────
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def export_teacher_stats(cls, teacher_list: list) -> io.BytesIO:
        """
        يُصدّر إحصائيات المراقبين إلى Excel.
        تنسيق مجمَّع: لكل مراقب رأس خاص به، يليه جدول تعييناته.
        """
        TEACHER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        INFO_FILL     = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        SUB_FILL      = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        ALT_FILL      = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        WHITE_FILL    = PatternFill(fill_type=None)

        wb = Workbook()
        ws = wb.active
        ws.title = "إحصائيات المراقبين"
        ws.sheet_view.rightToLeft = True

        # ── صف العنوان الرئيسي ──────────────────────────────────────
        ws.merge_cells("A1:D1")
        tc = ws["A1"]
        tc.value = "إحصائيات المراقبين"
        tc.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        tc.fill      = TEACHER_FILL
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border    = cls.THIN_BORDER
        ws.row_dimensions[1].height = 32

        current_row = 3  # blank row 2

        for teacher in teacher_list:
            assignments = getattr(teacher, "filtered_assignments", [])

            # ── رأس المراقب ──────────────────────────────────────────
            ws.merge_cells(f"A{current_row}:C{current_row}")
            nc = ws.cell(row=current_row, column=1, value=teacher.formatted_name)
            nc.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
            nc.fill      = INFO_FILL
            nc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            nc.border    = cls.THIN_BORDER

            cc = ws.cell(row=current_row, column=4,
                         value=f"عدد المرات: {teacher.total_count}")
            cc.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            cc.fill      = INFO_FILL
            cc.alignment = cls.CENTER
            cc.border    = cls.THIN_BORDER
            ws.row_dimensions[current_row].height = 26
            current_row += 1

            # ── صف معلومات اللقب والدرجة ─────────────────────────────
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ic = ws.cell(row=current_row, column=1,
                         value=f"اللقب: {teacher.title}    |    الدرجة: {teacher.degree}")
            ic.font      = Font(name="Arial", italic=True, size=10, color="1F4E79")
            ic.fill      = SUB_FILL
            ic.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            ic.border    = cls.THIN_BORDER
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            # ── رؤوس أعمدة التعيينات ─────────────────────────────────
            for col_i, hdr in enumerate(["التاريخ", "القاعة", "الوقت", ""], start=1):
                hc = ws.cell(row=current_row, column=col_i, value=hdr)
                hc.font      = Font(name="Arial", bold=True, size=10, color="1F4E79")
                hc.fill      = SUB_FILL
                hc.alignment = cls.CENTER
                hc.border    = cls.THIN_BORDER
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            # ── صفوف التعيينات ───────────────────────────────────────
            if assignments:
                for idx, asgn in enumerate(assignments):
                    row_fill = ALT_FILL if idx % 2 == 0 else WHITE_FILL
                    for col_i, val in enumerate(
                        [asgn.date, asgn.room_label, asgn.time, ""], start=1
                    ):
                        dc = ws.cell(row=current_row, column=col_i, value=val)
                        dc.font      = cls.CELL_FONT
                        dc.alignment = cls.CENTER
                        dc.border    = cls.THIN_BORDER
                        dc.fill      = row_fill
                    current_row += 1
            else:
                ws.merge_cells(f"A{current_row}:D{current_row}")
                ec = ws.cell(row=current_row, column=1, value="لا توجد بيانات")
                ec.font      = Font(name="Arial", italic=True, size=10, color="999999")
                ec.alignment = cls.CENTER
                ec.border    = cls.THIN_BORDER
                current_row += 1

            current_row += 1  # فاصل بين المراقبين

        # ── عرض الأعمدة ──────────────────────────────────────────────
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
