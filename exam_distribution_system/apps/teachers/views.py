import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from django.http import HttpResponse
from rest_framework import generics, filters, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Value, IntegerField
from django_filters.rest_framework import DjangoFilterBackend

from core.pagination import CustomPagination
from core.activity import log_activity
from .models import Teacher, TeacherExclusion
from .permissions import BaseTeacherPermission
from .serializers import (
    TeacherReadSerializer, TeacherWriteSerializer,
    TeacherExclusionReadSerializer, TeacherExclusionWriteSerializer,
)
from .services import delete_teacher

# ── Valid choice sets (mirrors model constants) ───────────────────────────────
_VALID_TITLES  = {"استاذ", "استاذ مساعد", "مدرس", "مدرس مساعد"}
_VALID_DEGREES = {"دكتوراه", "ماجستير", "بكالوريوس"}
_VALID_LANGS   = {"ar", "en"}


def _get_base_queryset():
    """
    Queryset مع annotation لـ distribution_count.
    في المرحلة الثالثة استبدل Value(0) بـ Count('tech_divides') الفعلية.
    """
    # مثال جاهز للمرحلة الثالثة:
    # from django.db.models import Count
    # return Teacher.objects.annotate(distribution_count=Count('tech_divides'))
    return Teacher.objects.annotate(
        distribution_count=Value(0, output_field=IntegerField())
    ).order_by("name")


class TeacherListCreateView(generics.ListCreateAPIView):
    """
    GET  /api/teachers/  - قائمة المدرّسين مع distribution_count
    POST /api/teachers/  - إضافة مدرّس جديد
    """

    permission_classes = [IsAuthenticated, BaseTeacherPermission]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["lang"]
    search_fields = ["name"]

    def get_queryset(self):
        return _get_base_queryset()

    def get_serializer_class(self):
        if self.request.method == "POST":
            return TeacherWriteSerializer
        return TeacherReadSerializer

    def create(self, request, *args, **kwargs):
        write_serializer = TeacherWriteSerializer(data=request.data)
        write_serializer.is_valid(raise_exception=True)
        instance = write_serializer.save(created_by_user_id=request.user.id)
        log_activity(
            user=request.user,
            action='create',
            module='teaching_management',
            description=f'إضافة مدرّس: {instance.name}',
            request=request,
        )
        # أعد قراءة الـ instance مع الـ annotation
        instance_with_annotation = _get_base_queryset().get(pk=instance.pk)
        read_serializer = TeacherReadSerializer(instance_with_annotation)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED)


class TeacherRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET    /api/teachers/{id}/
    PATCH  /api/teachers/{id}/
    DELETE /api/teachers/{id}/
    """

    permission_classes = [IsAuthenticated, BaseTeacherPermission]
    http_method_names = ["get", "patch", "delete", "head", "options"]

    def get_queryset(self):
        return _get_base_queryset()

    def get_serializer_class(self):
        if self.request.method == "PATCH":
            return TeacherWriteSerializer
        return TeacherReadSerializer

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        write_serializer = TeacherWriteSerializer(
            instance, data=request.data, partial=True
        )
        write_serializer.is_valid(raise_exception=True)
        updated = write_serializer.save()
        log_activity(
            user=request.user,
            action='update',
            module='teaching_management',
            description=f'تعديل مدرّس: {updated.name}',
            request=request,
        )
        # أعد قراءة مع annotation
        updated_with_annotation = _get_base_queryset().get(pk=updated.pk)
        return Response(TeacherReadSerializer(updated_with_annotation).data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        name = instance.name
        delete_teacher(instance)
        log_activity(
            user=request.user,
            action='delete',
            module='teaching_management',
            description=f'حذف مدرّس: {name}',
            request=request,
        )
        return Response(
            {"success": True, "message": "تم حذف المدرّس بنجاح."},
            status=status.HTTP_200_OK,
        )


# ─── Exclusion Views ──────────────────────────────────────────────────────────

class ExclusionListCreateView(generics.ListCreateAPIView):
    """
    GET  /api/teachers/exclusions/  — قائمة الاستثناءات (بحث + فلتر تاريخ)
    POST /api/teachers/exclusions/  — إضافة استثناء جديد
    """
    permission_classes = [IsAuthenticated, BaseTeacherPermission]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["teacher__name"]

    def get_queryset(self):
        qs = TeacherExclusion.objects.select_related("teacher").order_by("-date", "teacher__name")
        date = self.request.query_params.get("date", "").strip()
        if date:
            qs = qs.filter(date=date)
        return qs

    def get_serializer_class(self):
        if self.request.method == "POST":
            return TeacherExclusionWriteSerializer
        return TeacherExclusionReadSerializer

    def create(self, request, *args, **kwargs):
        write_ser = TeacherExclusionWriteSerializer(data=request.data)
        write_ser.is_valid(raise_exception=True)
        instance = write_ser.save(created_by_user_id=request.user.id)
        read_ser = TeacherExclusionReadSerializer(
            TeacherExclusion.objects.select_related("teacher").get(pk=instance.pk)
        )
        return Response(read_ser.data, status=status.HTTP_201_CREATED)


class ExclusionDestroyView(generics.DestroyAPIView):
    """DELETE /api/teachers/exclusions/{id}/ — حذف استثناء."""
    permission_classes = [IsAuthenticated, BaseTeacherPermission]
    queryset = TeacherExclusion.objects.all()

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return Response(
            {"success": True, "message": "تم حذف الاستثناء بنجاح."},
            status=status.HTTP_200_OK,
        )


# ─── Excel Import Views ───────────────────────────────────────────────────────

class TeacherTemplateDownloadView(APIView):
    """
    GET /api/teachers/import/template/
    يُنشئ ملف Excel جاهزاً للتعبئة بقوائم منسدلة وورقة تعليمات.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()

        # ── الألوان ──────────────────────────────────────────────────────────
        DARK_BLUE  = "1E3A5F"
        MID_BLUE   = "2563EB"
        LIGHT_BLUE = "EFF6FF"
        GOLD       = "F59E0B"
        LIGHT_GOLD = "FFFBEB"
        GREEN      = "16A34A"
        LIGHT_GREEN= "F0FDF4"
        GRAY_BG    = "F8FAFC"
        WHITE      = "FFFFFF"
        RED        = "DC2626"

        def _header_cell(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, size=11, bold=True, center=True):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(bold=bold, color=fg, name="Arial", size=size)
            c.alignment = Alignment(horizontal="center" if center else "right",
                                    vertical="center", wrap_text=True)
            return c

        def _data_cell(ws, row, col, value="", bg=WHITE, fg="000000", bold=False, center=False):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fg, name="Arial", size=10, bold=bold)
            c.alignment = Alignment(horizontal="center" if center else "right",
                                    vertical="center", wrap_text=True)
            return c

        # ════════════════════════════════════════════════════════════════════
        # ورقة 1 — البيانات
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "البيانات"
        ws1.sheet_view.rightToLeft = True
        ws1.sheet_view.showGridLines = True

        # ── صف العنوان ──────────────────────────────────────────────────────
        headers = [
            ("الاسم الكامل", 32),
            ("اللقب العلمي", 22),
            ("الشهادة العلمية", 20),
            ("اللغة (اختياري)", 18),
        ]
        for col_idx, (label, width) in enumerate(headers, start=1):
            _header_cell(ws1, 1, col_idx, label)
            ws1.column_dimensions[get_column_letter(col_idx)].width = width

        ws1.row_dimensions[1].height = 28

        # ── أمثلة ─────────────────────────────────────────────────────────
        SAMPLES = [
            ("أحمد محمد علي",    "استاذ",         "دكتوراه",  "ar"),
            ("سارة عبد الله",   "استاذ مساعد",   "دكتوراه",  "ar"),
            ("خالد إبراهيم",    "مدرس",          "ماجستير",  "ar"),
            ("منى حسن كريم",    "مدرس مساعد",    "بكالوريوس","ar"),
        ]
        for r_idx, sample in enumerate(SAMPLES, start=2):
            bg = LIGHT_BLUE if r_idx % 2 == 0 else WHITE
            for c_idx, val in enumerate(sample, start=1):
                _data_cell(ws1, r_idx, c_idx, val, bg=bg)
            ws1.row_dimensions[r_idx].height = 22

        # ── قوائم منسدلة (Data Validation) ──────────────────────────────────
        title_dv = DataValidation(
            type="list",
            formula1='"استاذ,استاذ مساعد,مدرس,مدرس مساعد"',
            allow_blank=False,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="قيمة غير صحيحة",
            error="اختر اللقب من القائمة: استاذ / استاذ مساعد / مدرس / مدرس مساعد",
        )
        degree_dv = DataValidation(
            type="list",
            formula1='"دكتوراه,ماجستير,بكالوريوس"',
            allow_blank=False,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="قيمة غير صحيحة",
            error="اختر الشهادة من القائمة: دكتوراه / ماجستير / بكالوريوس",
        )
        lang_dv = DataValidation(
            type="list",
            formula1='"ar,en"',
            allow_blank=True,
            showDropDown=False,
        )
        for dv, col in [(title_dv, "B"), (degree_dv, "C"), (lang_dv, "D")]:
            dv.sqref = f"{col}2:{col}2000"
            ws1.add_data_validation(dv)

        # ── تجميد الصف الأول ────────────────────────────────────────────────
        ws1.freeze_panes = "A2"

        # ════════════════════════════════════════════════════════════════════
        # ورقة 2 — تعليمات الاستخدام
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("تعليمات الاستخدام")
        ws2.sheet_view.rightToLeft = True
        ws2.sheet_view.showGridLines = False
        ws2.column_dimensions["A"].width = 2    # هامش
        ws2.column_dimensions["B"].width = 24
        ws2.column_dimensions["C"].width = 36
        ws2.column_dimensions["D"].width = 24
        ws2.column_dimensions["E"].width = 2    # هامش

        r = 1  # مؤشر الصف الحالي

        # ── عنوان رئيسي ─────────────────────────────────────────────────────
        ws2.merge_cells(f"B{r}:D{r}")
        c = ws2.cell(row=r, column=2,
                     value="📋  دليل استخدام ملف رفع بيانات المراقبين")
        c.fill  = PatternFill("solid", fgColor=DARK_BLUE)
        c.font  = Font(bold=True, color=WHITE, name="Arial", size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 36
        r += 1

        # ── فراغ ─────────────────────────────────────────────────────────────
        r += 1

        # ── الخطوات العامة ────────────────────────────────────────────────────
        def section_title(row, text):
            ws2.merge_cells(f"B{row}:D{row}")
            c = ws2.cell(row=row, column=2, value=text)
            c.fill  = PatternFill("solid", fgColor=MID_BLUE)
            c.font  = Font(bold=True, color=WHITE, name="Arial", size=11)
            c.alignment = Alignment(horizontal="right", vertical="center",
                                    indent=1)
            ws2.row_dimensions[row].height = 26

        def info_row(row, col_b, col_c, col_d="", bg=WHITE):
            for col, val in [(2, col_b), (3, col_c), (4, col_d)]:
                c = ws2.cell(row=row, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="right", vertical="center",
                                        wrap_text=True, indent=1)
            ws2.row_dimensions[row].height = 20

        section_title(r, "📌  خطوات الاستخدام")
        r += 1
        steps = [
            ("1", "انتقل إلى ورقة «البيانات»",
             "في نفس الملف — الورقة الأولى"),
            ("2", "احذف صفوف الأمثلة (السطور 2–5)",
             "البيانات الحمراء هي أمثلة توضيحية فقط"),
            ("3", "ابدأ الإدخال من الصف الثاني",
             "الصف الأول هو رأس الجدول، لا تمسه"),
            ("4", "استخدم القوائم المنسدلة",
             "اضغط على خلية اللقب أو الشهادة لتظهر الخيارات"),
            ("5", "احفظ الملف بصيغة .xlsx",
             'ملف ← حفظ باسم ← "Excel Workbook (*.xlsx)"'),
            ("6", "ارفع الملف في النظام",
             "اضغط «رفع ملف Excel» في صفحة التدريسيين"),
        ]
        for step_num, step_title, step_desc in steps:
            bg = LIGHT_BLUE if int(step_num) % 2 == 0 else WHITE
            info_row(r, f"  الخطوة {step_num}:  {step_title}", step_desc, bg=bg)
            r += 1

        r += 1

        # ── جدول القيم المسموح بها ────────────────────────────────────────────
        section_title(r, "📊  القيم المسموح بها لكل عمود")
        r += 1

        # رأس الجدول
        for col, label in [(2, "العمود"), (3, "القيم المقبولة"), (4, "ملاحظات")]:
            c = ws2.cell(row=r, column=col, value=label)
            c.fill = PatternFill("solid", fgColor="374151")
            c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 22
        r += 1

        table_data = [
            ("الاسم الكامل  (مطلوب)",
             "اكتب الاسم الكامل للمراقب",
             "لا حدود للطول — الاسم مطلوب دائماً",
             WHITE, LIGHT_BLUE),
            ("اللقب العلمي  (مطلوب)",
             "استاذ  /  استاذ مساعد  /  مدرس  /  مدرس مساعد",
             "يُحدد نوع المراقب في التوزيع (أستاذ أو مدرّس)",
             LIGHT_BLUE, LIGHT_BLUE),
            ("الشهادة العلمية  (مطلوب)",
             "دكتوراه  /  ماجستير  /  بكالوريوس",
             "تظهر مع الاسم في التقارير",
             WHITE, LIGHT_BLUE),
            ("اللغة  (اختياري)",
             "ar  (عربي)  /  en  (إنجليزي)",
             "القيمة الافتراضية: ar — إذا تركتها فارغة",
             LIGHT_BLUE, LIGHT_BLUE),
        ]
        for col_b, col_c, col_d, bg1, bg2 in table_data:
            for col, val, bg in [(2, col_b, bg1), (3, col_c, bg2), (4, col_d, bg1)]:
                c = ws2.cell(row=r, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="right", vertical="center",
                                        wrap_text=True, indent=1)
            ws2.row_dimensions[r].height = 30
            r += 1

        r += 1

        # ── جدول اللقب العلمي وتأثيره ────────────────────────────────────────
        section_title(r, "🎓  اللقب العلمي وأثره على التوزيع")
        r += 1
        for col, label in [(2, "اللقب العلمي"), (3, "نوع المراقب في التوزيع"), (4, "ترتيب الأولوية")]:
            c = ws2.cell(row=r, column=col, value=label)
            c.fill = PatternFill("solid", fgColor="374151")
            c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 22
        r += 1

        title_effect = [
            ("استاذ",         "أستاذ (النوع 2)",         "الأولوية الأعلى في التوزيع", LIGHT_GOLD),
            ("استاذ مساعد",  "أستاذ (النوع 2)",         "الأولوية الأعلى في التوزيع", LIGHT_GOLD),
            ("مدرس",          "مدرّس (النوع 1)",         "الأولوية الثانية",             LIGHT_GREEN),
            ("مدرس مساعد",   "مدرّس (النوع 1)",         "الأولوية الثانية",             LIGHT_GREEN),
        ]
        for title_val, t_type, priority, bg in title_effect:
            for col, val in [(2, title_val), (3, t_type), (4, priority)]:
                c = ws2.cell(row=r, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[r].height = 22
            r += 1

        r += 1

        # ── ملاحظات مهمة ─────────────────────────────────────────────────────
        section_title(r, "⚠️  ملاحظات مهمة")
        r += 1
        notes = [
            "لا تغيّر أسماء الأعمدة في الصف الأول — النظام يقرأها بترتيب ثابت.",
            "القيم في عمود «اللقب العلمي» و«الشهادة» يجب أن تكون بالعربية تماماً كما في القائمة.",
            "يمكنك إدخال مئات الأسماء في ورقة واحدة — لا يوجد حد أقصى.",
            "إذا كان هناك خطأ في صف، سيُتجاهل ذلك الصف وتُستكمل بقية الأسطر الصحيحة.",
            "النظام لا يتحقق من الأسماء المكررة — تأكد من عدم تكرار نفس المراقب.",
            "بعد الرفع، ستظهر نتيجة تفصيلية تُبيّن عدد من تم إدخاله والأخطاء إن وجدت.",
        ]
        for i, note in enumerate(notes):
            bg = LIGHT_BLUE if i % 2 == 0 else WHITE
            ws2.merge_cells(f"B{r}:D{r}")
            c = ws2.cell(row=r, column=2, value=f"  •  {note}")
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="right", vertical="center",
                                    wrap_text=True, indent=1)
            ws2.row_dimensions[r].height = 22
            r += 1

        # ─── تصدير ───────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="teachers_import_template.xlsx"'
        )
        return response


class TeacherImportView(APIView):
    """
    POST /api/teachers/import/
    يقرأ ملف Excel ويُنشئ المراقبين بشكل جماعي.
    يُعيد: { imported, failed, errors: [{row, name, errors}] }
    """
    permission_classes = [IsAuthenticated, BaseTeacherPermission]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "لم يتم إرسال أي ملف."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.lower().endswith((".xlsx", ".xls")):
            return Response(
                {"error": "نوع الملف غير مدعوم. يرجى رفع ملف Excel (.xlsx)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── قراءة الملف ────────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return Response(
                {"error": "تعذّر قراءة الملف. تأكد أن الملف بصيغة Excel صحيحة."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── التحقق من البيانات ─────────────────────────────────────────────
        teachers_to_create: list[Teacher] = []
        errors: list[dict] = []

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        for row_idx, row in enumerate(rows, start=2):
            # تجاهل الصفوف الفارغة
            if not row or not any(c for c in row if c is not None):
                continue

            name   = str(row[0]).strip() if row[0] is not None else ""
            title  = str(row[1]).strip() if row[1] is not None else ""
            degree = str(row[2]).strip() if row[2] is not None else ""
            lang   = str(row[3]).strip() if row[3] is not None else "ar"
            if lang not in _VALID_LANGS:
                lang = "ar"

            row_errors = []
            if not name:
                row_errors.append("الاسم مطلوب")
            if title not in _VALID_TITLES:
                row_errors.append(
                    f"اللقب «{title}» غير صحيح — "
                    "المقبول: استاذ / استاذ مساعد / مدرس / مدرس مساعد"
                )
            if degree not in _VALID_DEGREES:
                row_errors.append(
                    f"الشهادة «{degree}» غير صحيحة — "
                    "المقبول: دكتوراه / ماجستير / بكالوريوس"
                )

            if row_errors:
                errors.append({"row": row_idx, "name": name or f"صف {row_idx}",
                                "errors": row_errors})
            else:
                t = Teacher(
                    name=name, title=title, degree=degree, lang=lang,
                    created_by_user_id=request.user.id,
                )
                t.type = t._compute_type()   # احسب النوع قبل bulk_create
                teachers_to_create.append(t)

        if not teachers_to_create and not errors:
            return Response(
                {"error": "الملف لا يحتوي على بيانات."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── الإدخال الجماعي ────────────────────────────────────────────────
        created = Teacher.objects.bulk_create(teachers_to_create)

        log_activity(
            user=request.user,
            action="create",
            module="teaching_management",
            description=f"رفع {len(created)} مراقب عبر ملف Excel",
            request=request,
        )

        return Response(
            {
                "imported": len(created),
                "failed":   len(errors),
                "errors":   errors,
            },
            status=status.HTTP_200_OK,
        )
